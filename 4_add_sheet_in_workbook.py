# import openpyxl module
import openpyxl
  
# Call a Workbook() function of openpyxl 
# to create a new blank Workbook object
wb = openpyxl.Workbook()
  
sheet = wb.active
  
# Sheets can be added to workbook with the
# workbook object's create_sheet() method. 
wb.create_sheet(index = 1 , title = "demo sheet2")
  
wb.save("C:\\Users\\user\\Desktop\\demo.xlsx")
